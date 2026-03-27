#!/usr/bin/env python
# coding: utf-8

# In[1]:
#get_ipython().run_line_magic('pip', 'install pandas openpyxl matplotlib')

# In[2]:
import pandas as pd
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Border, Side

# Configuración inicial
directorio       = r"C:\Users\ffeli\Desktop\ProyectosPython\2- AsistenciaLab"
archivo_visitas   = os.path.join(directorio, "visitas_laboratorio.xlsx")
archivo_registros = os.path.join(directorio, "registros_laboratorio.xlsx")
archivo_salida    = os.path.join(directorio, "Resumen_Asistencia.xlsx")

def format_minutes_to_hhmm(minutos):
    horas = minutos // 60
    mins  = minutos % 60
    return f"{int(horas)}:{int(mins):02d}"

# Carga de datos
df_visitas   = pd.read_excel(archivo_visitas)
df_registros = pd.read_excel(archivo_registros)

df_visitas['RUT']   = df_visitas['RUT'].str.upper()
df_registros['RUT'] = df_registros['RUT'].str.upper()
df_visitas['Minutos'] = pd.to_numeric(df_visitas['Minutos'], errors='coerce').fillna(0)

df_visitas['Fecha'] = pd.to_datetime(df_visitas['Fecha'])
df_visitas['Mes']   = df_visitas['Fecha'].dt.month_name(locale='es_ES').str.capitalize()

meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

# Datos generales
total_alumnos       = df_registros['RUT'].nunique()
alumnos_por_carrera = df_registros.groupby('Carrera')['RUT'].nunique().reset_index(name='Cantidad de Alumnos')
alumnos_por_ramo    = df_registros.groupby('Asignatura')['RUT'].nunique().reset_index(name='Cantidad de Alumnos')

visitas_mensuales = df_visitas.groupby('Mes').size().reset_index(name='Cantidad de Visitas')
visitas_mensuales['Mes'] = pd.Categorical(visitas_mensuales['Mes'], categories=meses, ordered=True)
visitas_mensuales = visitas_mensuales.sort_values('Mes')



motivos_visita = df_visitas['Motivo'].value_counts().reset_index(name='Cantidad')
motivos_visita.columns = ['Motivo','Cantidad']

alumnos_nombres = df_registros[['RUT','Nombre','Apellido']].drop_duplicates()

# Detalle por alumno
detalle_alumno = df_registros[['RUT','Nombre','Apellido','Carrera','Asignatura']].drop_duplicates()

# Frecuencia por alumno
frecuencia = df_visitas.groupby(['RUT','Mes']).size().unstack(fill_value=0)
frecuencia['Total'] = frecuencia.sum(axis=1)
frecuencia = alumnos_nombres.merge(frecuencia, on='RUT', how='right')
meses_pres = [m for m in meses if m in frecuencia.columns]
frecuencia = frecuencia[['RUT','Nombre','Apellido'] + meses_pres + ['Total']]

# Horas y minutos por alumno
horas_minutos = df_visitas.groupby(['RUT','Mes'])['Minutos'].sum().unstack(fill_value=0)
horas_minutos['Total'] = horas_minutos.sum(axis=1)
horas_minutos = alumnos_nombres.merge(horas_minutos, on='RUT', how='right')
horas_minutos = horas_minutos[['RUT','Nombre','Apellido'] + meses_pres + ['Total']]

horas_fmt = horas_minutos.copy()
for col in meses_pres + ['Total']:
    horas_fmt[col] = horas_fmt[col].apply(format_minutes_to_hhmm)

# Exportar a Excel
with pd.ExcelWriter(archivo_salida, engine='openpyxl') as escritor:
    alumnos_por_carrera.to_excel(escritor, index=False, sheet_name='Por Carrera')
    alumnos_por_ramo.to_excel(escritor, index=False, sheet_name='Por Asignatura')
    detalle_alumno.to_excel(escritor, index=False, sheet_name='Detalle por Alumno')
    visitas_mensuales.to_excel(escritor, index=False, sheet_name='Visitas por Mes')
    motivos_visita.to_excel(escritor, index=False, sheet_name='Motivos de Visita')
    frecuencia.to_excel(escritor, index=False, sheet_name='Frecuencia por Alumno')
    horas_fmt.to_excel(escritor, index=False, sheet_name='Horas por Alumno')
    horas_minutos.to_excel(escritor, index=False, sheet_name='_Horas_num')

wb = load_workbook(archivo_salida)

# Totales por hoja
def fila_total(hoja, col):
    ws = wb[hoja]
    r  = ws.max_row + 1
    c  = chr(64 + col)
    ws.cell(row=r, column=1, value='TOTAL')
    ws.cell(row=r, column=col, value=f'=SUM({c}2:{c}{r-1})')

for h,col in [('Por Carrera',2),('Por Asignatura',2),
             ('Motivos de Visita',2),('Visitas por Mes',2)]:
    fila_total(h,col)

# Descripción en frecuencia
ws_f = wb['Frecuencia por Alumno']
ws_f.insert_rows(1)
ws_f['A1'] = "Esta hoja muestra la cantidad de visitas por alumno, mes a mes, con total al final."
ws_f['A1'].font = Font(bold=True)

# Resumen General
if 'Resumen General' in wb.sheetnames:
    wb.remove(wb['Resumen General'])
ws_res = wb.create_sheet('Resumen General')

total_minutos = df_visitas['Minutos'].sum()
resumen = [
    ("Total de Alumnos Registrados", total_alumnos),
    ("Total de Carreras Distintas", alumnos_por_carrera.shape[0]),
    ("Total de Asignaturas Distintas", alumnos_por_ramo.shape[0]),
    ("Total de Motivos de Visita Distintos", motivos_visita.shape[0]),
    ("Total de Registros de Visita", df_visitas.shape[0]),
    ("Total de Horas en Laboratorio", format_minutes_to_hhmm(total_minutos))
]
for i,(d,v) in enumerate(resumen, start=1):
    ws_res.cell(row=i, column=1, value=d).font = Font(bold=True)
    ws_res.cell(row=i, column=2, value=v)
ws_res.column_dimensions['A'].width = 45
ws_res.column_dimensions['B'].width = 20

# Gráfico de barras horizontal
plt.figure(figsize=(10,6))
oc = alumnos_por_carrera.sort_values('Cantidad de Alumnos')
plt.barh(oc['Carrera'], oc['Cantidad de Alumnos'])
for i,v in enumerate(oc['Cantidad de Alumnos']):
    plt.text(v+1, i, str(v), va='center')
plt.title('Alumnos Registrados por Carrera')
plt.tight_layout()
p1 = os.path.join(directorio,'grafico_carrera.png')
plt.savefig(p1); plt.close()

plt.figure(figsize=(10,5))
plt.plot(visitas_mensuales['Mes'], visitas_mensuales['Cantidad de Visitas'], marker='o')
for i,v in enumerate(visitas_mensuales['Cantidad de Visitas']):
    plt.text(i, v+1, str(v), ha='center', fontsize=9)
plt.title('Visitas al Laboratorio por mes')
plt.tight_layout()
p2 = os.path.join(directorio,'grafico_visitas.png')
plt.savefig(p2); plt.close()

plt.figure(figsize=(8,6))
plt.pie(motivos_visita['Cantidad'], labels=motivos_visita['Motivo'], autopct='%1.1f%%', startangle=140)
plt.title('Distribución de Motivos de Visita')
plt.tight_layout()
p3 = os.path.join(directorio,'grafico_motivos.png')
plt.savefig(p3); plt.close()

detalle = df_registros[['RUT','Nombre','Apellido','Carrera','Asignatura']].drop_duplicates()
pivot = detalle.pivot_table(index='Carrera', columns='Asignatura', values='RUT', aggfunc=pd.Series.nunique, fill_value=0)
fig,ax = plt.subplots(figsize=(12,8))
pivot.plot(kind='bar', ax=ax)
for p in ax.patches:
    ax.annotate(str(int(p.get_height())), (p.get_x()+p.get_width()/2, p.get_height()), ha='center', va='bottom', fontsize=8)
plt.title('Distribución de Alumnos por Carrera en cada Asignatura')
plt.ylabel('Cantidad de Alumnos')
plt.tight_layout()
p4 = os.path.join(directorio,'grafico_detalle.png')
plt.savefig(p4); plt.close()

ws_g = wb.create_sheet('Gráficos')
for img,pos in [(p1,'A1'),(p2,'A35'),(p3,'S1'),(p4,'S35')]:
    im = XLImage(img); im.anchor = pos; ws_g.add_image(im)

lado = Side(border_style='thin', color='000000')
borde = Border(top=lado, bottom=lado, left=lado, right=lado)
for nombre in wb.sheetnames:
    if nombre!='Gráficos':
        ws = wb[nombre]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = borde

wb['_Horas_num'].sheet_state = 'hidden'
wb.save(archivo_salida)

for f in [p1,p2,p3,p4]:
    try: os.remove(f)
    except: pass

